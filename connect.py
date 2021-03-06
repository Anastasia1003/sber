from jira import JIRA
import openpyxl
from openpyxl.styles import Font

user = 'rogulinaanastasia@gmail.com'
apikey = '**'  #token
server = 'https://practice-sber2020.atlassian.net'

options = {
    'server': server
}
jira = JIRA(options, basic_auth=(user, apikey))

# выводит задачи по созданным epics
jql = 'project= TeamProject AND issuetype = Task' # project= TeamProject AND parent in (TEAM-1, TEAM-59, TEAM-32)
issues = jira.search_issues(jql, maxResults=100)

for issue in issues:
    print('{}: {}'.format(issue.key, issue.fields.summary))

filepath = "/Users/apple/Documents/requests.xlsx"  # указываем путь для файла выгрузки результатов поиска в excel
wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row=1, column=1)
c1.value = "Название проекта"
c1.font = Font(size=14, underline='single', bold=True, italic=False)
c2 = sheet.cell(row=2, column=1)
c2.value = "TeamProject"
c3 = sheet.cell(row=1, column=2)
c3.value = "Количество задач"
c3.font = Font(size=14, underline='single', bold=True, italic=False)
c4 = sheet.cell(row=2, column=2)
c4.value = len(issues)
wb.create_sheet(index=1, title="request2")
wb.active = 1
jql = 'project= TeamProject AND issuetype = Task AND status = Реализовано'  # выводит задачи со статусом "Реализовано"
issues = jira.search_issues(jql)

for sheet in wb:
    if sheet.title == 'request2':
        sheet.sheet_view.tabSelected = True
    else:
        sheet.sheet_view.tabSelected = False

c5 = sheet.cell(row=1, column=1)
c5.value = "Ключ"
c5.font = Font(size=14, underline='single', bold=True, italic=False)
c6 = sheet.cell(row=1, column=2)
c6.value = "Задача"
c6.font = Font(size=14, underline='single', bold=True, italic=False)
c6 = sheet.cell(row=1, column=3)
c6.value = "Статус"
c6.font = Font(size=14, underline='single', bold=True, italic=False)

for issue, i in zip(issues, range(2, len(issues) + 2)):
    print(issue.key)
    id_task = sheet.cell(row=i, column=1)
    id_task.value = issue.key
    task_name = sheet.cell(row=i, column=2)
    task_name.value = issue.fields.summary
    status = sheet.cell(row=i, column=3)
    status.value = "Реализовано"
wb.active = 0
wb.save(filepath)
